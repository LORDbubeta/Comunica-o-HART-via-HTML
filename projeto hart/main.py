from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse  # Adicionando o import de HTMLResponse
import serial
import time
import struct
from openpyxl import Workbook, load_workbook
import os
from pydantic import BaseModel
from pymodbus.client import ModbusTcpClient



# Configuração do CLP Delta
PLC_IP = "192.168.1.5"  # IP do seu CLP
PLC_PORT = 502          # Porta Modbus padrão
MEMORY_M40_ADDRESS = 40  # Endereço Modbus da memória M15
REGISTRO_D306_ADDRESS = 306 # Endereço Modbus do registro D306
REGISTRO_D123_ADDRESS = 123 # Endereço Modbus do registro D123

# Criar cliente Modbus
plc = ModbusTcpClient(PLC_IP, port=PLC_PORT)

app = FastAPI()

SERIAL_PORT_1 = "COM5"  # Altere para a porta correta do seu dispositivo
SERIAL_PORT = "COM6"  # Altere para a porta correta do seu dispositivo
BAUDRATE = 1200
MAX_RETRIES = 5  # Número máximo de tentativas para conexão
RETRY_INTERVAL = 5  # Intervalo entre tentativas de reconexão em segundos

EXCEL_FILE = "dados_hart.xlsx"


# Função para tentar conectar ao dispositivo HART
def connect_to_device():
    attempt = 0
    while attempt < MAX_RETRIES:
        try:
            ser = serial.Serial(
                port=SERIAL_PORT,
                baudrate=BAUDRATE,
                parity=serial.PARITY_ODD,
                stopbits=serial.STOPBITS_ONE,
                bytesize=serial.EIGHTBITS,
                timeout=1
            )
            print(f"Conectado ao dispositivo HART na porta {SERIAL_PORT}")
            return ser
        except serial.SerialException as e:
            attempt += 1
            print(f"Tentativa {attempt} de conexão falhou: {str(e)}")
            if attempt < MAX_RETRIES:
                print(f"Tentando novamente em {RETRY_INTERVAL} segundos...")
                time.sleep(RETRY_INTERVAL)
            else:
                raise HTTPException(status_code=500, detail="Falha ao conectar ao dispositivo HART após várias tentativas.")


# Função para tentar conectar ao dispositivo HART
def connect_to_device_1():
    attempt = 0
    while attempt < MAX_RETRIES:
        try:
            ser_1 = serial.Serial(
                    port=SERIAL_PORT_1,
                    baudrate=BAUDRATE,
                    parity=serial.PARITY_ODD,
                    stopbits=serial.STOPBITS_ONE,
                    bytesize=serial.EIGHTBITS,
                    timeout=1
            )
            print(f"Conectado ao dispositivo HART na porta {SERIAL_PORT_1}")
            return ser_1
        except serial.SerialException as e:
            attempt += 1
            print(f"Tentativa {attempt} de conexão falhou: {str(e)}")
            if attempt < MAX_RETRIES:
                print(f"Tentando novamente em {RETRY_INTERVAL} segundos...")
                time.sleep(RETRY_INTERVAL)
            else:
                raise HTTPException(status_code=500, detail="Falha ao conectar ao dispositivo HART após várias tentativas.")

# Tenta conectar ao dispositivo HART assim que o servidor for iniciado
ser_1 = connect_to_device_1()

# Tenta conectar ao dispositivo HART assim que o servidor for iniciado
ser = connect_to_device()


final_command = None

final_command_1 = None

resultados = []

def float_to_ieee754_32bit(value):
    """
    Converte um valor float para o formato IEEE 754 de 32 bits (ponto flutuante).
    """
    # Converte o float para bytes usando o formato 'f' (ponto flutuante de 4 bytes)
    packed = struct.pack('!f', value)  # '!f' significa Big Endian e tipo float
    # Converte os bytes para hexadecimal
    return ''.join(f'{byte:02X}' for byte in packed)

def calculate_checksum(command):
    lrc = 0
    for byte in command:  # Itera sobre cada byte da mensagem
        lrc ^= byte  # Aplica a operação XOR com cada byte
    return lrc


def save_to_excel(data):
    # Se o arquivo não existir, criar um novo
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados HART"
        # Criar cabeçalhos
        ws.append(["Código Resposta", "Status", "Código Unidade", "Descrição Unidade", "Valor_float"])
        wb.save(EXCEL_FILE)

    # Abrir o arquivo e adicionar uma nova linha de dados
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data["codigo_resposta"],
        data["status_e"],
        data["codigo_unidade"],
        data["descricao_unidade"],
        data["valor_float"]
    ])
    wb.save(EXCEL_FILE)

def build_command_1(device_id, manufacturer_id, binario_id):

    global final_command_1
    
    # Base do comando
    base_command = 'FFFFFFFFFF'

    # Transformar o bit 1 após "fe" (mudar o bit 1 para "10")
    transformed = format(int(binario_id, 16), '08b')  # Converter o binário corretamente
    transformed = '1' + transformed[1:]  # Colocar "1" à esquerda
    transformed_hex = format(int(transformed, 2), '02x')  # Converter de volta para hexadecimal
    
    command_part = '82' + transformed_hex  # Adicionar o número modificado ao comando

    # ID do fabricante (bit 2 após FE)
    manufacturer_bit = format(int(manufacturer_id, 16), '02x')  # Converter o ID do fabricante para hex
    command_part += manufacturer_bit
    
    # Convertendo o device_id para hexadecimal e preenchendo os bits 9, 10 e 11
    device_id_hex = format(int(device_id, 16), '06x')  # 6 caracteres hexadecimais para o ID do dispositivo
    command_part += device_id_hex + '01' + '00'  # ID do dispositivo
    
    # Converter o comando em bytes e calcular o checksum
    command = bytes.fromhex(command_part)
    checksum = calculate_checksum(command)
    checksum_hex = format(checksum, '02x')
    
    # Comando base modificado
    final_command_base = base_command + command_part
    print(final_command_base)
    
    # Comando final com o checksum
    final_command_1 = final_command_base + checksum_hex
    print(final_command_1)
    return final_command_1


def build_command(device_id, manufacturer_id, binario_id):

    global final_command
    
    # Base do comando
    base_command = 'FFFFFFFFFF'

    # Transformar o bit 1 após "fe" (mudar o bit 1 para "10")
    transformed = format(int(binario_id, 16), '08b')  # Converter o binário corretamente
    transformed = '1' + transformed[1:]  # Colocar "1" à esquerda
    transformed_hex = format(int(transformed, 2), '02x')  # Converter de volta para hexadecimal
    
    command_part = '82' + transformed_hex  # Adicionar o número modificado ao comando

    # ID do fabricante (bit 2 após FE)
    manufacturer_bit = format(int(manufacturer_id, 16), '02x')  # Converter o ID do fabricante para hex
    command_part += manufacturer_bit
    
    # Convertendo o device_id para hexadecimal e preenchendo os bits 9, 10 e 11
    device_id_hex = format(int(device_id, 16), '06x')  # 6 caracteres hexadecimais para o ID do dispositivo
    command_part += device_id_hex + '01' + '00'  # ID do dispositivo
    
    # Converter o comando em bytes e calcular o checksum
    command = bytes.fromhex(command_part)
    checksum = calculate_checksum(command)
    checksum_hex = format(checksum, '02x')
    
    # Comando base modificado
    final_command_base = base_command + command_part
    print(final_command_base)
    
    # Comando final com o checksum
    final_command = final_command_base + checksum_hex
    print(final_command)
    return final_command


def process_hart_response(response_hex):
    # Remover os primeiros 4 bytes "FF"
    response_hex = response_hex.lstrip("Ff")
    
    # Converter para bytes
    response_bytes = bytes.fromhex(response_hex)
    
    # Verificar se a resposta tem o tamanho mínimo esperado
    if len(response_bytes) < 14:
        return {"erro": "Resposta muito curta"}

    # Localizar os valores na string hexadecimal
    response_code_hex = response_hex[16:18]  # Após "07" e antes de "40"
    response_code = response_bytes[8]  # Byte 7
    print(response_code)
    if response_code != 0x00:
        response_error_codes = {
            0xC0: "Erro de comunicação: paridade no frame recebido",
            0xA0: "Erro de comunicação: overrun. Byte sobrescrito",
            0x90: "Erro de comunicação: framing. Stop bit não detectado",
            0x88: "Erro de comunicação: checksum calculado",
            0x82: "Erro de comunicação: overflow. Frame muito longo",
            0x02: "Erro de comando: seleção inválida",
            0x05: "Erro de comando: poucos dados recebidos",
            0x07: "Erro de comando: escrita protegida",
        }
        return {"erro": response_error_codes.get(response_code, "Erro desconhecido")}
    
    # Status (Após "00" e antes de "0A")
    status_code_hex = response_hex[18:20]  # Pegando o hexadecimal correto
    status_code = response_bytes[9]
    status_definitions = {
        0x80: "Mau funcionamento do equipamento",
        0x40: "Configuração alterada",
        0x20: "Reinicialização do equipamento",
        0x10: "Status adicionais disponíveis",
        0x08: "Corrente de saída em modo fixo",
        0x04: "Corrente de saída saturada",
        0x02: "Variável não-primária fora dos limites",
        0x01: "Variável primária fora dos limites",
    }
    status_message = status_definitions.get(status_code, "Status desconhecido")
    
    # Unidade (Após "40" e antes de "B9")
    unit_code_hex = response_hex[20:22]
    unit_code = response_bytes[10]  # Converter para decimal
    unit_codes = {
        1: "inH2O (68 °F)", 2: "inHg (0 °C)", 3: "ftH2O (68 °F)",
        4: "mmH2O (68 °F)", 5: "mmHg (0 °C)", 6: "psi", 7: "bar",
        8: "mbar", 9: "g/cm²", 10: "kg/cm²", 11: "Pa", 12: "kPa",
        13: "Torr", 14: "atm", 15: "ft³/min", 16: "gal/min", 17: "l/min",
        18: "gal(UK)/min", 19: "m³/h", 20: "gal/s", 23: "Mgal/d",
        24: "l/s", 25: "Ml/d", 26: "ft³/s", 27: "ft³/d", 28: "m³/s",
        29: "m³/d", 30: "gal(UK)/h", 31: "gal(UK)/d", 32: "Nm³/h",
        123: "SCF/min", 124: "ft³/h", 131: "m³/min", 132: "bbl/min",
        133: "bbl/h", 134: "bbl/d", 135: "gal/h", 136: "gal(UK)/s",
        137: "gal(UK)/h", 170: "cmH2O (4 °C)", 171: "mH2O (4 °C)",
        172: "cmHg (0 °C)", 173: "lb/ft²", 174: "hPa", 175: "psia",
        176: "kg/m²", 177: "ftH2O (4 °C)", 178: "ftH2O (60 °F)",
        179: "mHg (0 °C)", 180: "Mpsi", 181: "oz/in²", 237: "MPa",
        238: "inH2O (4 °C)", 239: "mmH2O (4 °C)"
    }
    unit_description = unit_codes.get(unit_code, "Unidade desconhecida")

    # Valor IEEE-754 (4 bytes após unidade)
    float_hex = response_hex[22:30]  # 4 bytes em hexadecimal
    float_bytes = bytes.fromhex(float_hex)  # Converter para bytes
    float_bin = ''.join(format(byte, '08b') for byte in float_bytes)  # Converter para binário
    print(float_bin)
    print(float_hex)
    
    if len(float_bytes) == 4:
        float_value = struct.unpack('>f', float_bytes)[0]  # IEEE-754
    else:
        float_value = None
    
    return {
        "codigo_resposta": response_code_hex,  # Hexadecimal
        "status_e": status_message,
        "codigo_unidade": unit_code_hex,  # Hexadecimal
        "descricao_unidade": unit_description,
        "valor_hex": float_hex,  # Valor hexadecimal
        "valor_binario": float_bin,  # Binário antes da conversão IEEE-754
        "valor_float": float_value,  # Número final convertido
    }

@app.get("/html", response_class=HTMLResponse)
def get_html_page():
    return """
    <!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Comando HART</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #f4f4f9;
            color: #333;
            padding: 20px;
        }
        h1 {
            color: #4CAF50;
        }
        input, select, button {
            padding: 10px;
            font-size: 16px;
            margin: 10px 0;
            border-radius: 5px;
        }
        button {
            background-color: #4CAF50;
            color: white;
            border: none;
            cursor: pointer;
        }
        button:hover {
            background-color: #45a049;
        }
        .result {
            margin-top: 20px;
            padding: 15px;
            border: 1px solid #ddd;
            background-color: #fff;
            border-radius: 5px;
        }
    </style>
</head>
<body>
    <h1>Comando HART</h1>
    
    <div>
        <label for="hart_1">Endereço do Primeiro Dispositivo:</label>
        <input type="text" id="hart_1" placeholder="Coloque o endereço do dispositivo">
        <button onclick="montarComando(1)">Enviar Comando</button>
    </div>
    <div id="result1" class="result">Aguardando comando...</div>

    <div>
        <label for="hart_2">Endereço do Segundo Dispositivo:</label>
        <input type="text" id="hart_2" placeholder="Coloque o endereço do dispositivo">
        <button onclick="montarComando(2)">Enviar Comando</button>
    </div>
    <div id="result2" class="result">Aguardando comando...</div>

    <div>
        <button onclick="coletarValores(); coleta_clp()">Coletar Valores</button>
        <div id="resultColeta" class="result">Aguardando coleta de valores...</div>
    </div>

    <div>
        <button onclick="comando_zero()">Trim Zero</button>
        <div id="zero" class="result">Aguardando zero...</div>
    </div>

    <div>
        <label for="padrao">Escolha o padrão:</label>
        <select id="padrao">
            <option value="1">BAR 0 á 10</option>
            <option value="2">BAR 0 á 30</option>
            <option value="3">BAR 0 á 100</option>
            <option value="4">BAR 0 á 250</option>
        </select>
    </div>
    
    <div>
        <label for="pressao">Escolha a unidade de pressão:</label>
        <select id="pressao">
            <option value="7">Bar</option>
            <option value="10">Kgf/cm²</option>
            <option value="4">mmH₂O</option>
            <option value="5">mmHg</option>
            <option value="6">PSI</option>
            <option value="8">mBar</option>
        </select>
    </div>
    
    <div>
        <label for="range">Digite o range máximo:</label>
        <input type="number" id="range">
    </div>

    <button type="button" onclick="transferir()">Passar valores</button>

    <script>
        function montarComando(dispositivo) {
            const endereco = document.getElementById(`hart_${dispositivo}`).value;
            if (!endereco) {
                alert("Por favor, insira o endereço do dispositivo.");
                return;
            }
            fetch(`/dispositivo_${dispositivo}`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ endereco })
            })
            .then(response => response.json())
            .then(data => {
                document.getElementById(`result${dispositivo}`).textContent = `Resultado: ${data.message}`;
            })
            .catch(error => {
                console.error('Erro na requisição:', error);
                document.getElementById(`result${dispositivo}`).textContent = 'Erro ao detectar dispositivo.';
            });
        }

function coletarValores() {
    fetch('/comando_1')
        .then(response => response.json())
        .then(data => {
            console.log(data);  // Verificar resposta JSON no console

            if (data.resultados && Array.isArray(data.resultados)) {
                let resultado_1 = data.resultados.find(item => item.dispositivo_1);
                let resultado_2 = data.resultados.find(item => item.dispositivo_2);

                if (!resultado_1 || !resultado_2) {
                    document.getElementById('resultColeta').textContent = 'Erro: Estrutura da resposta incorreta.';
                    return;
                }

                resultado_1 = resultado_1.dispositivo_1;
                resultado_2 = resultado_2.dispositivo_2;

                const resultText = `
                    <strong>Dispositivo 1:</strong><br>
                    Valor: ${resultado_1.valor_float !== undefined ? resultado_1.valor_float.toFixed(4) : "N/A"} ${resultado_1.descricao_unidade || "Unidade desconhecida"}<br>
                    Status: ${resultado_1.status_e || "Sem status"}<br>
                    Código Unidade: ${resultado_1.codigo_unidade || "N/A"}<br><br>

                    <strong>Dispositivo 2:</strong><br>
                    Valor: ${resultado_2.valor_float !== undefined ? resultado_2.valor_float.toFixed(4) : "N/A"} ${resultado_2.descricao_unidade || "Unidade desconhecida"}<br>
                    Status: ${resultado_2.status_e || "Sem status"}<br>
                    Código Unidade: ${resultado_2.codigo_unidade || "N/A"}<br><br>
                `;

                document.getElementById('resultColeta').innerHTML = resultText;

            } else {
                document.getElementById('resultColeta').textContent = 'Erro: Estrutura da resposta incorreta.';
            }
        })

               
       
        .catch(error => {
            console.error('Erro na requisição:', error);
            document.getElementById('resultColeta').textContent = 'Erro ao coletar os dados.';
        });
}

    function coleta_clp() {
    fetch('/coleta_clp')  // Chama o endpoint do backend
        .then(response => response.json())  // Converte a resposta para JSON
        .then(data => {
            if (data.status === 'ativo') {
                console.log("Coleta feita no CLP");
            } else {
                console.log(data.mensagem || "Erro desconhecido");
            }
        })
        .catch(error => {
            console.error('Erro na requisição:', error);
        });
}

      function comando_zero() {
        fetch('/comando_setzero')
            .then(response => {
                if (!response.ok) {
                    throw new Error('Erro na requisição');
                }
                return response.json(); // Converte a resposta em JSON
            })
            .then(data => {
                // Aqui você pode manipular os dados retornados
                console.log('Resposta do servidor:', data);
                // Exemplo: Exibindo a resposta no frontend
                document.getElementById('zero').textContent = 'Comando setzero executado com sucesso!';
            })
            .catch(error => {
                console.error('Erro ao coletar valores:', error);
                document.getElementById('zero').textContent = 'Erro ao coletar valores dos dispositivos.';
            });
    }

       function transferir() {
       const rangeValue = parseFloat(document.getElementById('range').value);  // Garante que o valor seja um número float
       const pressValue = parseInt(document.getElementById('pressao').value);  // Garante que o valor seja um número inteiro
       const padraoValue = parseInt(document.getElementById('padrao').value);  // Garante que o valor seja um número inteiro

       // Verificar se os valores são válidos
       if (isNaN(rangeValue) || isNaN(padraoValue) || isNaN(pressValue)) {
           alert("Por favor, insira valores válidos para o range, pressão e padrão.");
           return;
       }

        if (!rangeValue || !pressValue || !padraoValue) {
        alert("Por favor, preencha todos os campos.");
        return; // Não envia a requisição se algum campo estiver vazio
      }

        // Exibir os valores no console para garantir que estão corretos
        console.log("Range:", rangeValue, "Pressão:", pressValue, "Padrão:", padraoValue);

        

        const payload = {
            rangee: rangeValue,
            pressao: pressValue,
            padrao: padraoValue
        };

        console.log("Dados enviados:", payload);  // Verificar os dados no console


        // Enviar os dados para o backend
        fetch('/escrever_memoria', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify(payload)
        })
        .then(response => response.json())
        .then(data => {
            // Verificar o status da resposta
            if (data.status === "escrito") {
                alert("Dados escritos com sucesso!");
            } else {
                alert(`Erro ao escrever os dados: ${data.mensagem}`);
            }
        })
        .catch(error => {
            console.error('Erro ao transferir valores:', error);
            alert('Erro ao escrever valores no dispositivo.');
        });
    }
        
    </script>
</body>
</html>
    """


class DeviceData(BaseModel):
    rangee: float
    pressao: int
    padrao: int


@app.post("/escrever_memoria")
async def escrever_memoria(data: DeviceData):

    print("Range:", data.rangee)
    print("Pressão:", data.pressao)
    print("Padrão:", data.padrao)

    range_max = int(data.rangee)

    try:
        # Conectar ao PLC
        if not plc.connect():
            return {"status": "erro", "mensagem": "Não foi possível conectar ao CLP"}

        # Escrever no registro de range
        response_1 = plc.write_register(REGISTRO_D306_ADDRESS, range_max)
        if response_1.isError():
            return {"status": "erro", "mensagem": "Erro ao escrever valor no registro D306"}

        # Escrever no registro de padrão
        response_2 = plc.write_register(REGISTRO_D123_ADDRESS, data.padrao )
        if response_2.isError():
            return {"status": "erro", "mensagem": "Erro ao escrever valor no registro D307"}

        # Se necessário, executar o comando especial de pressão
        comando_especial(data)

        return {"status": "escrito", "mensagem": "Todos os valores foram escritos com sucesso!"}

    except Exception as e:
        return {"status": "erro", "mensagem": f"Ocorreu um erro: {str(e)}"}


def comando_especial(data: DeviceData):

    global final_command_1
    print(final_command_1)
    
    # Implementação do comando especial baseado nos dados de pressão e range
    pressao = data.pressao
    range_value = data.rangee
    print(f"Comando especial com pressão: {pressao} e range: {range_value}")
    
    #unidade
    # Implementação do comando especial baseado nos dados de unidade pressão

    base_comando_unit = final_command_1[0:10]
    comando_unit = final_command_1[10:22]
    print(comando_unit)
    
    ler_press_part = comando_unit + '2C' + '01' #pressão vai no meio do codigo tem que transformar para bits

    pressao_hex = hex(pressao)[2:].zfill(2).upper()  # Converte para hexadecimal e garante 2 caracteres

    ler_press_part += pressao_hex

    print(ler_press_part)

    check_comando = calculate_checksum(bytes.fromhex(ler_press_part))
    
    print(check_comando)

    check_hex = format(check_comando, '02x')

    ler_press = base_comando_unit + ler_press_part + check_hex

    ler_press_bytes = bytes.fromhex(ler_press)
    ser_1.write(ler_press_bytes)
    time.sleep(2)
    press_resultado = ser_1.read(128)
    print("Resposta recebida da porta 2:", press_resultado.hex())

    time.sleep(10)
    
    #range
    # Implementação do comando especial baseado nos dados de unidade pressão

    base_comando_range = final_command_1[0:10]
    comando_range = final_command_1[10:22]
    print(comando_range)
    
    ler_range_part = comando_range + '23' + '09' #range vai no meio do codigo tem que transformar para bits

    ler_range_part += pressao_hex

    range_hex = float_to_ieee754_32bit(range_value)

    print(range_hex)
    
    ler_range_part += range_hex + '00' + '00' + '00' + '00'

    print(ler_range_part)

    check_comando_2 = calculate_checksum(bytes.fromhex(ler_range_part))
    
    check_hex_2 = format(check_comando_2, '02x')

    ler_range = base_comando_range + ler_range_part + check_hex_2

    ler_range_bytes = bytes.fromhex(ler_range)
    ser_1.write(ler_range_bytes)
    time.sleep(2)
    range_resultado = ser_1.read(128)
    print("Resposta recebida da porta 2:", range_resultado.hex())

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws['J7'] = range_value
    wb.save(EXCEL_FILE)

@app.get("/comando_setzero")
def comando_zero():

    global final_command_1
    print(final_command_1)
    
    #zero
    # Implementação do comando especial baseado nos dados de unidade pressão

    base_comando_zero = final_command_1[0:10]
    comando_zero = final_command_1[10:22]
    print(comando_zero)
    
    ler_zero_part = comando_zero + '2B' + '00' #range vai no meio do codigo tem que transformar para bits

    print(ler_zero_part)

    check_comando_3 = calculate_checksum(bytes.fromhex(ler_zero_part))

    check_hex_3 = format(check_comando_3, '02x')

    set_zero = base_comando_zero + ler_zero_part + check_hex_3

    set_zero_bytes = bytes.fromhex(set_zero)
    ser_1.write(set_zero_bytes)
    time.sleep(2)
    zero_resultado = ser_1.read(128)
    print("Resposta recebida da porta 2:", zero_resultado.hex())

    # Retornando uma resposta JSON para o frontend
    return {"status": "success", "message": "Comando setzero executado com sucesso", "resultado": zero_resultado.hex()}
    

    
@app.get("/comando_1")
def comand_1():
    global final_command_1  # Comando da primeira porta
    global final_command  # Comando da segunda porta
    print(final_command_1)
    print(final_command)
    global resultados

    resultados.clear()
    
    print("Enviando comando (hex) para dispositivo 1:", final_command)
    ser.write(bytes.fromhex(final_command))
    response_1 = ser.read(128)
    print("Resposta recebida da porta 1:", response_1.hex())

    resultado_1 = process_hart_response(response_1.hex())

    save_to_excel(resultado_1)
    print(resultado_1)

    resultados.append({"dispositivo_1": resultado_1})  # Alterado para dicionário correto


    print("Enviando comando (hex) para dispositivo 2:", final_command_1)
    ser_1.write(bytes.fromhex(final_command_1))
    response_2 = ser_1.read(128)
    print("Resposta recebida da porta 2:", response_2.hex())

    resultado_2 = process_hart_response(response_2.hex())

    save_to_excel(resultado_2)
    print(resultado_2)

    resultados.append({"dispositivo_2": resultado_2})

    print(resultados)

    coleta_clp()

    return {"resultados": resultados}  # Retorna os resultados corretamente

@app.get("/coleta_clp")
def coleta_clp():

    try:
        # Conexão com o CLP e lógica de coleta
        if not plc.connect():
            return {"status": "erro", "mensagem": "Não foi possível conectar ao CLP"}

        plc.write_coils(MEMORY_M40_ADDRESS, [True])
        response = plc.read_coils(MEMORY_M40_ADDRESS)

        print(response)

        if response.isError():
            return {"status": "erro", "mensagem": "Erro ao ler memória M40"}

        if response.bits[0]:  # Se o bit M40 estiver ativado, começa a coleta
            return {"status": "ativo", "mensagem": "Coleta foi ativada com sucesso no CLP"}

        return {"status": "erro", "mensagem": "Coleta não foi ativada no CLP"}

    except Exception as e:
        return {"status": "erro", "mensagem": f"Erro desconhecido: {str(e)}"}
        


def command_0_0(endereco_hart1: str):    
    global final_command  # Tornar a variável global

    base_command = 'FFFFFFFFFF'

    command_part = '02'

    command_part += endereco_hart1  # Usando o valor recebido no frontend

    command_part += '00' + '00'

    command_check = bytes.fromhex(command_part)
    checksum = calculate_checksum(command_check)
    checksum_hex = format(checksum, '02x')

    command = bytes.fromhex(base_command + command_part + checksum_hex)
    print(f"Comando gerado (em hexadecimal): {command.hex()}")

    try:
        # Envia o comando HART
        ser.write(command)
        time.sleep(1)  # Aguarda um pouco para o dispositivo responder
        
        # Lê a resposta (128 bytes é um tamanho seguro para a maioria das respostas)
        response = ser.read(128)
        
        # Converter a resposta para hexadecimal
        response_hex = response.hex()
        print(f"Resposta recebida (em hexadecimal): {response_hex}")

        # Ignora os primeiros 8 bits "FF" (tanto minúsculo quanto maiúsculo)
        # Aqui, verificamos e removemos os "FF" iniciais, se existirem
        response_hex = response_hex.lstrip('fF')  # Remove 'f' ou 'F' do começo
        
        # Verifica se a resposta tem dados suficientes após remover os FF
        if len(response_hex) == "":
            raise ValueError("Resposta incompleta ou inválida após remover os FF.")

        # Pegando o ID do dispositivo da resposta
        device_id = response_hex[30:36]  # Extraímos 8 caracteres hexadecimais a partir do byte 16
        print(f"ID do dispositivo extraído: {device_id}")

        # Pegando o ID do fabricante da resposta
        manufacturer_id = response_hex[16:18]  # Extraímos 2 caracteres hexadecimais a partir do byte 12
        print(f"ID do fabricante extraído: {manufacturer_id}")

        # Pegando o binário ID (depois do manufacturer_id)
        binario_id = response_hex[14:16]  # Extraímos 2 caracteres hexadecimais
        print(f"Binário ID extraído: {binario_id}")

        # Chamando a função de construção de comando com os IDs extraídos
        final_command = build_command(device_id, manufacturer_id, binario_id)
        
        print(f"Comando gerado para o dispositivo: {final_command}")
        
        # Exibe o comando final em formato hexadecimal
        return {"status": "sucesso", "response": final_command}
        

    except Exception as e:
        print(f"Erro ao enviar comando: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao enviar comando: {str(e)}")
    

def command_0_1(endereco_hart2: str):    
    global final_command_1  # Tornar a variável global

    base_command = 'FFFFFFFFFF'

    command_part = '02'

    command_part += endereco_hart2  # Usando o valor recebido no frontend

    command_part += '00' + '00'

    command_check = bytes.fromhex(command_part)
    checksum = calculate_checksum(command_check)
    checksum_hex = format(checksum, '02x')

    command = bytes.fromhex(base_command + command_part + checksum_hex)
    print(f"Comando gerado (em hexadecimal): {command.hex()}")

    try:
        # Envia o comando HART
        ser_1.write(command)
        time.sleep(1)  # Aguarda um pouco para o dispositivo responder
        
        # Lê a resposta (128 bytes é um tamanho seguro para a maioria das respostas)
        response = ser_1.read(128)
        
        # Converter a resposta para hexadecimal
        response_hex = response.hex()
        print(f"Resposta recebida (em hexadecimal): {response_hex}")

        # Ignora os primeiros 8 bits "FF" (tanto minúsculo quanto maiúsculo)
        # Aqui, verificamos e removemos os "FF" iniciais, se existirem
        response_hex = response_hex.lstrip('fF')  # Remove 'f' ou 'F' do começo
        
        # Verifica se a resposta tem dados suficientes após remover os FF
        if len(response_hex) == "":
            raise ValueError("Resposta incompleta ou inválida após remover os FF.")

        # Pegando o ID do dispositivo da resposta
        device_id = response_hex[30:36]  # Extraímos 8 caracteres hexadecimais a partir do byte 16
        print(f"ID do dispositivo extraído: {device_id}")

        # Pegando o ID do fabricante da resposta
        manufacturer_id = response_hex[16:18]  # Extraímos 2 caracteres hexadecimais a partir do byte 12
        print(f"ID do fabricante extraído: {manufacturer_id}")

        # Pegando o binário ID (depois do manufacturer_id)
        binario_id = response_hex[14:16]  # Extraímos 2 caracteres hexadecimais
        print(f"Binário ID extraído: {binario_id}")

        # Chamando a função de construção de comando com os IDs extraídos
        final_command_1 = build_command_1(device_id, manufacturer_id, binario_id)
        
        print(f"Comando gerado para o dispositivo: {final_command_1}")
        
        # Exibe o comando final em formato hexadecimal
        return {"status": "sucesso", "response": final_command_1}

    except Exception as e:
        print(f"Erro ao enviar comando: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao enviar comando: {str(e)}")
    

# Defina um modelo de dados para receber a requisição JSON
class DeviceData(BaseModel):
    endereco: str

# Exemplo de resposta do comando HART
response_example = "FFFFFFFF86A62329FDF0010700400ab9500f48c5"
result = process_hart_response(response_example)
print(result)

@app.post("/dispositivo_1")
async def command_hart1(device_data: DeviceData):
    # Agora o "endereco" do dispositivo estará disponível na variável
    endereco_hart1 = device_data.endereco
    print(f"Endereço HART 1 recebido: {endereco_hart1}")
    command_0_0(endereco_hart1)

    # Agora você chama a função de montagem de comando, passando o endereço
    return {"message": f"Dispositivo 1 com o endereço {endereco_hart1} detectado!"}

@app.post("/dispositivo_2")
async def command_hart2(device_data: DeviceData):
    # Agora o "endereco" do dispositivo estará disponível na variável
    endereco_hart2 = device_data.endereco
    print(f"Endereço HART 2 recebido: {endereco_hart2}")
    command_0_1(endereco_hart2)

    # Agora você chama a função de montagem de comando, passando o endereço
    return {"message": f"Dispositivo 2 com o endereço {endereco_hart2} detectado!"}
    

@app.get("/")
def root():
    return {"message": "API HART rodando"}
